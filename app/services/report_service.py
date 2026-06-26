import io
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas
from app.models import db, Driver, Vehicle, Customer, Booking, Payment, Expense

class NumberedCanvas(canvas.Canvas):
    """Canvas implementation to compute total page numbers and draw running page footers."""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            super().showPage()
        super().save()

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 9)
        self.setFillColor(colors.HexColor("#6b7280"))
        
        # Header
        self.drawString(54, 750, "Jolly Cabs Operations Management Suite (JOMS) - Confidential")
        self.setStrokeColor(colors.HexColor("#e5e7eb"))
        self.setLineWidth(0.5)
        self.line(54, 742, 558, 742)
        
        # Footer
        page_text = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(558, 40, page_text)
        self.drawString(54, 40, f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self.line(54, 52, 558, 52)
        
        self.restoreState()


class ReportService:
    """Extracts business logs and compiles professional CSV, Excel, and PDF reports."""
    
    @staticmethod
    def get_report_data(report_type):
        """Prepares standard headers and data lists for the specified report category."""
        headers = []
        rows = []
        title = ""
        
        if report_type == 'revenue':
            title = "Revenue and Payments Report"
            headers = ["Payment ID", "Booking ID", "Customer", "Amount (INR)", "Method", "Status", "Date"]
            payments = Payment.query.join(Booking).join(Customer).order_by(Payment.payment_date.desc()).all()
            for p in payments:
                rows.append([
                    p.id,
                    p.booking_id,
                    p.booking.customer.name,
                    f"{p.amount:.2f}",
                    p.payment_method,
                    p.status,
                    p.payment_date.strftime('%Y-%m-%d %H:%M')
                ])
                
        elif report_type == 'driver':
            title = "Drivers Compliance and Performance Report"
            headers = ["Driver ID", "Driver Name", "Phone", "License No", "License Expiry", "Availability", "Status", "Total Trips"]
            drivers = Driver.query.all()
            for d in drivers:
                trips_count = Booking.query.filter_by(driver_id=d.id).count()
                rows.append([
                    d.id,
                    d.name,
                    d.phone,
                    d.license_number,
                    d.license_expiry.strftime('%Y-%m-%d'),
                    d.availability,
                    d.status,
                    trips_count
                ])
                
        elif report_type == 'vehicle':
            title = "Vehicle Fleet Status Report"
            headers = ["Vehicle ID", "Model", "Number Plate", "Fuel Type", "Insurance Expiry", "Permit Expiry", "Maintenance", "Status"]
            vehicles = Vehicle.query.all()
            for v in vehicles:
                rows.append([
                    v.id,
                    v.model,
                    v.vehicle_number,
                    v.fuel_type,
                    v.insurance_expiry.strftime('%Y-%m-%d'),
                    v.permit_expiry.strftime('%Y-%m-%d'),
                    v.maintenance_status,
                    v.availability
                ])
                
        elif report_type == 'customer':
            title = "Registered Customers Report"
            headers = ["Cust ID", "Name", "Phone", "Email", "Address", "Status", "Total Trips", "Total Spend (INR)"]
            customers = Customer.query.all()
            for c in customers:
                bookings = Booking.query.filter_by(customer_id=c.id).all()
                total_spent = sum(b.fare for b in bookings if b.payment_status == 'Paid')
                rows.append([
                    c.id,
                    c.name,
                    c.phone,
                    c.email,
                    c.address,
                    c.status,
                    len(bookings),
                    f"{total_spent:.2f}"
                ])
                
        elif report_type == 'bookings':
            title = "Bookings Ledger"
            headers = ["Book ID", "Customer", "Driver", "Vehicle", "Route (Pick - Drop)", "Fare (INR)", "Trip Status", "Payment"]
            bookings = Booking.query.order_by(Booking.booking_time.desc()).all()
            for b in bookings:
                c_name = b.customer.name
                d_name = b.driver.name if b.driver else "Unassigned"
                v_num = b.vehicle.vehicle_number if b.vehicle else "Unassigned"
                route = f"{b.pickup_location} -> {b.drop_location}"
                rows.append([
                    b.id,
                    c_name,
                    d_name,
                    v_num,
                    route,
                    f"{b.fare:.2f}",
                    b.status,
                    b.payment_status
                ])
                
        return title, headers, rows

    @staticmethod
    def generate_csv(report_type):
        """Generates a CSV byte stream for client download."""
        _, headers, rows = ReportService.get_report_data(report_type)
        
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        
        # Convert String Buffer to Bytes for Flask attachment response
        bytes_out = io.BytesIO()
        bytes_out.write(output.getvalue().encode('utf-8'))
        bytes_out.seek(0)
        return bytes_out

    @staticmethod
    def generate_excel(report_type):
        """Generates a styled Excel workbook byte stream."""
        title, headers, rows = ReportService.get_report_data(report_type)
        
        wb = Workbook()
        ws = wb.active
        ws.title = report_type.capitalize()[:30] # Limit tab name
        
        # Styles
        title_font = Font(name='Segoe UI', size=16, bold=True, color='1E3A8A')
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
        data_font = Font(name='Segoe UI', size=10)
        
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )
        
        # Write Title
        ws.append([title])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(row=1, column=1).font = title_font
        ws.row_dimensions[1].height = 40
        ws.cell(row=1, column=1).alignment = Alignment(vertical='center')
        
        ws.append([]) # Empty spacer row
        
        # Write Headers
        ws.append(headers)
        header_row_idx = 3
        ws.row_dimensions[header_row_idx].height = 25
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
        # Write Data
        for row in rows:
            ws.append(row)
            curr_row = ws.max_row
            ws.row_dimensions[curr_row].height = 20
            for col_idx in range(1, len(row) + 1):
                cell = ws.cell(row=curr_row, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
                
        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.row == 1:
                    continue  # Ignore title row length
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        excel_stream = io.BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)
        return excel_stream

    @staticmethod
    def generate_pdf(report_type):
        """Generates a professional PDF report with table formats and headers."""
        title, headers, rows = ReportService.get_report_data(report_type)
        
        pdf_stream = io.BytesIO()
        # Letter page size, margin 0.75 in (54 pt)
        doc = SimpleDocTemplate(
            pdf_stream,
            pagesize=letter,
            rightMargin=54,
            leftMargin=54,
            topMargin=72,
            bottomMargin=72
        )
        
        styles = getSampleStyleSheet()
        
        # Custom Typography Styles
        title_style = ParagraphStyle(
            name='DocTitle',
            fontName='Helvetica-Bold',
            fontSize=22,
            leading=26,
            textColor=colors.HexColor("#1e3a8a"), # Deep Blue
            spaceAfter=20
        )
        
        cell_header_style = ParagraphStyle(
            name='CellHeader',
            fontName='Helvetica-Bold',
            fontSize=9,
            leading=11,
            textColor=colors.white,
            alignment=1 # Centered
        )
        
        cell_data_style = ParagraphStyle(
            name='CellData',
            fontName='Helvetica',
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#1f2937") # Charcoal
        )
        
        story = []
        
        # Document Title
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 10))
        
        # Prepare Data Table
        # We wrap values in Paragraphs to enable text wrap inside reportlab table cells
        table_data = []
        
        # Headers Row
        header_row = [Paragraph(h, cell_header_style) for h in headers]
        table_data.append(header_row)
        
        # Data Rows
        for r in rows:
            table_row = [Paragraph(str(val), cell_data_style) for val in r]
            table_data.append(table_row)
            
        # Column Widths dynamically fitted (Letter width = 612 pt, margins = 108 pt. Printable width = 504 pt)
        col_count = len(headers)
        col_width = 504.0 / col_count
        widths = [col_width] * col_count
        
        # Let's adjust widths a bit for bookings table to give more space to Route
        if report_type == 'bookings' and col_count == 8:
            # col index 4 is route
            widths = [40, 60, 60, 60, 134, 50, 50, 50]
        elif report_type == 'customer' and col_count == 8:
            # col index 4 is address
            widths = [30, 60, 60, 80, 134, 40, 50, 50]
            
        t = Table(table_data, colWidths=widths, repeatRows=1)
        
        # Style the Table
        t_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
        ])
        
        # Alternating row colors
        for i in range(1, len(rows) + 1):
            if i % 2 == 0:
                t_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor("#f8fafc"))
                
        t.setStyle(t_style)
        story.append(t)
        
        # Build Document
        doc.build(story, canvasmaker=NumberedCanvas)
        
        pdf_stream.seek(0)
        return pdf_stream
